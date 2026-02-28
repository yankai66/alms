"""
资产出入门工单管理API
提供创建、查询、处理资产出入门工单的功能
"""

from fastapi import APIRouter, Depends, HTTPException, Query, Body, Path, UploadFile, File, Form
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, and_, func, cast, String
from typing import Optional, Dict, Any, List
from datetime import datetime
from pathlib import Path as FilePath
import httpx
import json
import uuid
import os

from app.db.session import get_db
from app.models.asset_models import Asset, WorkOrder, WorkOrderItem
from app.schemas.asset_schemas import ApiResponse, ResponseCode
from app.schemas.asset_entry_exit_schemas import (
    AssetEntryExitWorkOrderCreate,
    AssetEntryExitWorkOrderResponse,
    AssetEntryExitWorkOrderQuery,
    AssetEntryExitWorkOrderProcess
)
from app.core.config import settings
from app.core.logging_config import get_logger
from app.constants.operation_types import OperationType, OperationResult

router = APIRouter()
logger = get_logger(__name__)


# =====================================================
# 文件上传辅助函数
# =====================================================

def get_file_extension(filename: str) -> str:
    """获取文件扩展名（小写）"""
    if '.' in filename:
        return filename.rsplit('.', 1)[1].lower()
    return ''


def generate_filename(original_filename: str, prefix: str = "") -> str:
    """生成唯一文件名"""
    ext = get_file_extension(original_filename)
    date_str = datetime.now().strftime("%Y%m%d")
    unique_id = uuid.uuid4().hex[:8]
    
    if prefix:
        return f"{prefix}_{date_str}_{unique_id}.{ext}"
    return f"{date_str}_{unique_id}.{ext}"


def ensure_upload_dir(subdir: str = "") -> FilePath:
    """确保上传目录存在"""
    base_dir = FilePath(settings.PICTURE_DIR)
    if subdir:
        upload_dir = base_dir / subdir
    else:
        upload_dir = base_dir
    
    upload_dir.mkdir(parents=True, exist_ok=True)
    return upload_dir


# =====================================================
# 工单系统集成
# =====================================================

async def create_external_entry_exit_work_order(
    work_order_data: AssetEntryExitWorkOrderCreate,
    batch_id: str,
    creator_name: str = "system"
) -> Dict[str, Any]:
    """
    调用外部工单系统创建资产出入门工单
    """
    try:
        priority_value = str(work_order_data.priority)
        business_type_value = str(work_order_data.business_type)
        entry_exit_type_value = str(work_order_data.entry_exit_type)
        entry_exit_scope_value = str(work_order_data.entry_exit_scope)
        device_sn_text = ", ".join(work_order_data.device_sns)
        
        # 出入范围映射
        scope_labels = {
            "datacenter": "出入机房",
            "campus": "出入园区",
            "internal": "机房园区内出入"
        }
        scope_label = scope_labels.get(entry_exit_scope_value, entry_exit_scope_value)
        
        # 出入类型映射
        type_labels = {
            "move_in": "搬入",
            "move_out": "搬出"
        }
        type_label = type_labels.get(entry_exit_type_value, entry_exit_type_value)
        
        # 构建描述信息
        description_parts = [
            f"出入类型: {type_label}",
            f"出入范围: {scope_label}",
            f"出入原因: {work_order_data.entry_exit_reason}",
            f"出入日期: {work_order_data.entry_exit_date}",
            f"业务类型: {business_type_value}",
            f"机房: {work_order_data.datacenter}",
            f"优先级: {priority_value}",
            f"设备SN: {device_sn_text}",
            f"服务内容: {work_order_data.service_content}",
        ]
        
        if work_order_data.source_order_number:
            description_parts.append(f"来源单号: {work_order_data.source_order_number}")
        if work_order_data.remark:
            description_parts.append(f"备注: {work_order_data.remark}")
        if work_order_data.attachments:
            description_parts.append(f"附件数量: {len(work_order_data.attachments)}")
        
        # 构建variables
        variables = {
            "assignee": work_order_data.assignee
        }
        
        # 构建metadata（简化版）
        metadata = {
            "orderType": "asset_accounting",  # 资产出入门
            "assignee": work_order_data.assignee
        }
        
        work_order_request = {
            "title": work_order_data.title,
            "description": "\n".join(description_parts),
            "secretInfo": "11111",
            "creator": settings.WORK_ORDER_CREATOR,
            "creatorName": creator_name,
            "processId": "AssetEntryExit",
            "variables": variables,
            "externalBizId": batch_id,
            "bussinessMetaData": metadata
        }
        
        # 发送HTTP请求
        timeout = httpx.Timeout(60.0, connect=30.0)
        async with httpx.AsyncClient(timeout=timeout, follow_redirects=True) as client:
            print(f"[资产出入门工单创建] 正在连接工单系统: {settings.WORK_ORDER_API_URL}")
            print(f"[资产出入门工单创建] 请求报文: {json.dumps(work_order_request, ensure_ascii=False)}")
            
            headers = {
                "appid": settings.WORK_ORDER_APPID,
                "username": settings.WORK_ORDER_USERNAME,
                "Content-Type": "application/json"
            }
            if settings.WORK_ORDER_COOKIE:
                headers["Cookie"] = settings.WORK_ORDER_COOKIE

            response = await client.post(
                settings.WORK_ORDER_API_URL,
                headers=headers,
                json=work_order_request
            )
            
            response.raise_for_status()
            result = response.json()
            print(f"[资产出入门工单创建] 外部系统响应: {json.dumps(result, ensure_ascii=False)}")

            if result.get("status") != 0:
                error_msg = result.get("msg") or result.get("message") or "工单系统返回失败"
                print(f"[资产出入门工单创建失败] 状态码: {result.get('status')}, 错误信息: {error_msg}")
                return {
                    "success": False,
                    "error": error_msg,
                    "data": result
                }

            work_order_number = None
            if result.get("data"):
                work_order_number = result.get("data", {}).get("order_number")

            return {
                "success": True,
                "data": result,
                "work_order_number": work_order_number
            }
            
    except httpx.HTTPStatusError as e:
        error_msg = f"工单系统返回错误: {e.response.status_code} - {e.response.text}"
        print(f"[资产出入门工单创建失败] {error_msg}")
        return {
            "success": False,
            "error": error_msg,
            "status_code": e.response.status_code
        }
    except httpx.RequestError as e:
        error_detail = str(e) if str(e) else f"连接失败: {type(e).__name__}"
        error_msg = f"工单系统请求失败: {error_detail}"
        print(f"[资产出入门工单创建失败] {error_msg}")
        return {
            "success": False,
            "error": error_msg,
            "url": settings.WORK_ORDER_API_URL
        }
    except Exception as e:
        error_msg = f"工单创建异常: {str(e)}"
        print(f"[资产出入门工单创建失败] {error_msg}")
        return {
            "success": False,
            "error": error_msg
        }


# =====================================================
# 工单管理接口
# =====================================================

@router.post("/create", summary="创建资产出入门工单",
             response_model=ApiResponse,
             responses={
                 200: {
                     "description": "工单创建成功",
                     "content": {
                         "application/json": {
                             "example": {
                                 "code": 0,
                                 "message": "工单创建成功",
                                 "data": {
                                     "batch_id": "AEE_20251210120000",
                                     "work_order_number": "WO202512101234",
                                     "operation_type": "asset_accounting",
                                     "title": "服务器设备搬入",
                                     "status": "pending",
                                     "work_order_status": "processing",
                                     "datacenter": "DC01",
                                     "device_count": 2,
                                     "device_sns": ["SN123456", "SN789012"],
                                     "priority": "normal",
                                     "business_type": "other",
                                     "service_content": "新采购服务器搬入机房上架",
                                     "entry_exit_type": "move_in",
                                     "entry_exit_scope": "datacenter",
                                     "entry_exit_reason": "新设备采购到货，需搬入机房进行上架部署",
                                     "entry_exit_date": "2025-12-10",
                                     "sign_date": "2025-12-10",
                                     "assignee": "张三",
                                     "creator_name": "李四",
                                     "remark": "请提前准备好机柜空间",
                                     "attachments": ["https://download.sihua.tech/alms/images/asset_entry_exit/entry_exit_20251211_a1b2c3d4.jpg"],
                                     "created_at": "2025-12-10T12:00:00"
                                 }
                             }
                         }
                     }
                 },
                 404: {"description": "设备SN不存在"},
                 500: {"description": "服务器内部错误"}
             })
async def create_asset_entry_exit_work_order(
    title: str = Form(..., description="工单标题"),
    device_sns: str = Form(..., description="设备SN列表，逗号分隔"),
    entry_exit_type: str = Form(..., description="出入类型：move_in-搬入, move_out-搬出"),
    entry_exit_scope: str = Form(..., description="出入范围：datacenter-出入机房, campus-出入园区, internal-机房园区内出入"),
    entry_exit_reason: str = Form(..., description="出入原因"),
    assignee: str = Form(..., description="指派人"),
    creator_name: str = Form(..., description="创建人姓名"),
    datacenter: Optional[str] = Form(None, description="机房"),
    priority: Optional[str] = Form(None, description="优先级：normal-一般, urgent-紧急"),
    business_type: Optional[str] = Form(None, description="业务类型：fault_support-故障支持, change_support-变更支持, other-其他"),
    service_content: Optional[str] = Form(None, description="服务内容"),
    entry_exit_date: Optional[str] = Form(None, description="出入日期（YYYY-MM-DD格式）"),
    remark: Optional[str] = Form(None, description="备注"),
    source_order_number: Optional[str] = Form(None, description="来源单号"),
    campus_auth_order_number: Optional[str] = Form(None, description="园区授权单号"),
    campus_auth_status: Optional[str] = Form(None, description="园区授权状态"),
    device_type: Optional[str] = Form(None, description="设备类型"),
    sign_date: Optional[str] = Form(None, description="签字日期（YYYY-MM-DD格式）"),
    attachments: List[UploadFile] = File(default=[], description="附件文件列表（支持图片和文档，最多10个）"),
    db: Session = Depends(get_db)
):
    """
    创建资产出入门工单（支持附件上传）
    
    ## 功能说明
    用于创建资产搬入或搬出机房/园区的工单，支持同时上传附件文件。
    
    ## 请求格式
    Content-Type: multipart/form-data
    
    ## 必填字段 (7个)
    - **title**: 工单标题
    - **device_sns**: 设备SN列表（逗号分隔）
    - **entry_exit_type**: 出入类型（move_in-搬入, move_out-搬出）
    - **entry_exit_scope**: 出入范围（datacenter-出入机房, campus-出入园区, internal-机房园区内出入）
    - **entry_exit_reason**: 出入原因
    - **assignee**: 指派人
    - **creator_name**: 创建人姓名
    
    ## 可选字段
    - **datacenter**: 机房
    - **priority**: 优先级（normal-一般, urgent-紧急）
    - **business_type**: 业务类型（fault_support-故障支持, change_support-变更支持, other-其他）
    - **service_content**: 服务内容
    - **entry_exit_date**: 出入日期（格式: YYYY-MM-DD）
    - **remark**: 备注
    - **source_order_number**: 来源单号
    - **sign_date**: 签字日期（格式: YYYY-MM-DD）
    - **attachments**: 附件文件（最多10个，支持jpg/png/gif/pdf/doc/xlsx等）
    
    ## curl示例
    ```bash
    curl -X POST "http://localhost:8000/api/v1/asset-entry-exit-work-order/create" \\
      -F "title=服务器设备搬入" \\
      -F "device_sns=SN123456,SN789012" \\
      -F "entry_exit_type=move_in" \\
      -F "entry_exit_scope=datacenter" \\
      -F "entry_exit_reason=新设备采购到货" \\
      -F "assignee=张三" \\
      -F "creator_name=李四" \\
      -F "attachments=@photo1.jpg" \\
      -F "attachments=@photo2.jpg"
    ```
    """
    try:
        # 1. 解析设备SN列表
        device_sn_list = [sn.strip() for sn in device_sns.split(',') if sn.strip()]
        if not device_sn_list:
            return ApiResponse(
                code=ResponseCode.PARAM_ERROR,
                message="设备SN列表不能为空",
                data=None
            )
        
        # 2. 验证设备是否存在
        missing_sns = []
        existing_assets = {}
        
        for sn in device_sn_list:
            asset = db.query(Asset).filter(Asset.serial_number == sn).first()
            if not asset:
                missing_sns.append(sn)
            else:
                existing_assets[sn] = asset
        
        if missing_sns:
            return ApiResponse(
                code=ResponseCode.NOT_FOUND,
                message=f"以下设备SN不存在: {', '.join(missing_sns[:10])}{'...' if len(missing_sns) > 10 else ''}",
                data={"missing_sns": missing_sns}
            )
        
        # 3. 处理附件上传
        attachment_urls = []
        if attachments:
            # 过滤掉空文件
            valid_attachments = [f for f in attachments if f.filename]
            
            if len(valid_attachments) > 10:
                return ApiResponse(
                    code=ResponseCode.BAD_REQUEST,
                    message="一次最多上传10个附件",
                    data=None
                )
            
            allowed_extensions = {'jpg', 'jpeg', 'png', 'gif', 'bmp', 'webp', 'pdf', 'doc', 'docx', 'xls', 'xlsx'}
            
            for file in valid_attachments:
                ext = get_file_extension(file.filename)
                if ext not in allowed_extensions:
                    return ApiResponse(
                        code=ResponseCode.BAD_REQUEST,
                        message=f"不支持的文件类型: {ext}，支持的类型: {', '.join(allowed_extensions)}",
                        data=None
                    )
                
                content = await file.read()
                
                if len(content) > 10 * 1024 * 1024:
                    return ApiResponse(
                        code=ResponseCode.BAD_REQUEST,
                        message=f"文件 {file.filename} 过大，最大允许 10MB",
                        data=None
                    )
                
                filename = generate_filename(file.filename, "entry_exit")
                upload_dir = ensure_upload_dir("asset_entry_exit")
                file_path = upload_dir / filename
                
                with open(file_path, "wb") as f:
                    f.write(content)
                
                url = f"{settings.PICTURE_HTTP.rstrip('/')}/asset_entry_exit/{filename}"
                attachment_urls.append(url)
        
        # 4. 构建工单数据对象
        work_order_data = AssetEntryExitWorkOrderCreate(
            title=title,
            device_sns=device_sn_list,
            entry_exit_type=entry_exit_type,
            entry_exit_scope=entry_exit_scope,
            entry_exit_reason=entry_exit_reason,
            assignee=assignee,
            creator_name=creator_name,
            datacenter=datacenter,
            priority=priority,
            business_type=business_type,
            service_content=service_content,
            entry_exit_date=entry_exit_date,
            remark=remark,
            source_order_number=source_order_number,
            campus_auth_order_number=campus_auth_order_number,
            campus_auth_status=campus_auth_status,
            device_type=device_type,
            sign_date=sign_date,
            attachments=attachment_urls if attachment_urls else None
        )
        
        # 2. 生成批次ID (使用AEE前缀: Asset Entry Exit)
        batch_id = f"AEE_{datetime.now().strftime('%Y%m%d%H%M%S')}"
        
        # 3. 创建外部工单
        creator_name = work_order_data.creator_name or "system"
        work_order_result = await create_external_entry_exit_work_order(
            work_order_data,
            batch_id,
            creator_name
        )
        
        if not work_order_result.get("success"):
            logger.error("资产出入门工单创建失败", extra={
                "operationObject": work_order_data.title,
                "operationType": OperationType.ASSET_ENTRY_EXIT_CREATE,
                "operator": creator_name,
                "result": OperationResult.FAILED,
                "operationDetail": f"机房: {work_order_data.datacenter}, 出入类型: {work_order_data.entry_exit_type}, 出入范围: {work_order_data.entry_exit_scope}, 错误: {work_order_result.get('error', '未知错误')}"
            })
            return ApiResponse(
                code=ResponseCode.INTERNAL_ERROR,
                message=work_order_result.get("error", "工单创建失败"),
                data=work_order_result
            )
        
        # 4. 保存工单到本地数据库
        work_order_number = work_order_result.get("work_order_number")
        
        # 构建extra字段
        extra_data = {
            "priority": str(work_order_data.priority),
            "business_type": str(work_order_data.business_type),
            "device_sns": work_order_data.device_sns,
            "service_content": work_order_data.service_content,
            "entry_exit_type": str(work_order_data.entry_exit_type),
            "entry_exit_scope": str(work_order_data.entry_exit_scope),
            "entry_exit_reason": work_order_data.entry_exit_reason,
            "entry_exit_date": work_order_data.entry_exit_date,
            "creator_name": work_order_data.creator_name,
            "campus_auth_order_number": campus_auth_order_number,
            "campus_auth_status": campus_auth_status,
            "device_type": device_type,
            "sign_date": work_order_data.sign_date,
            "attachments": work_order_data.attachments or []
        }
        
        if work_order_data.source_order_number:
            extra_data["source_order_number"] = work_order_data.source_order_number
        
        # 创建本地工单记录
        scope_labels = {"datacenter": "出入机房", "campus": "出入园区", "internal": "机房园区内出入"}
        type_labels = {"move_in": "搬入", "move_out": "搬出"}
        
        description_text = (
            f"出入类型: {type_labels.get(str(work_order_data.entry_exit_type), work_order_data.entry_exit_type)}\n"
            f"出入范围: {scope_labels.get(str(work_order_data.entry_exit_scope), work_order_data.entry_exit_scope)}\n"
            f"出入原因: {work_order_data.entry_exit_reason}\n"
            f"出入日期: {work_order_data.entry_exit_date}\n"
            f"机房: {work_order_data.datacenter}\n"
            f"设备数量: {len(work_order_data.device_sns)}"
        )
        
        local_work_order = WorkOrder(
            batch_id=batch_id,
            work_order_number=work_order_number,
            operation_type="asset_accounting",
            title=work_order_data.title,
            description=description_text,
            status="pending",
            work_order_status="processing",  # 外部工单状态：创建成功即为进行中
            creator=creator_name,
            assignee=work_order_data.assignee,
            datacenter=work_order_data.datacenter,
            source_order_number=work_order_data.source_order_number,
            device_count=len(work_order_data.device_sns),
            extra=extra_data,
            remark=work_order_data.remark
        )
        
        try:
            db.add(local_work_order)
            db.flush()
            
            # 5. 创建工单明细
            for sn in work_order_data.device_sns:
                asset = existing_assets[sn]
                
                operation_data = {
                    "serial_number": sn,
                    "asset_tag": asset.asset_tag,
                    "asset_name": asset.name,
                    "datacenter": work_order_data.datacenter,
                    "priority": str(work_order_data.priority),
                    "business_type": str(work_order_data.business_type),
                    "service_content": work_order_data.service_content,
                    "entry_exit_type": str(work_order_data.entry_exit_type),
                    "entry_exit_scope": str(work_order_data.entry_exit_scope),
                    "entry_exit_reason": work_order_data.entry_exit_reason,
                    "entry_exit_date": work_order_data.entry_exit_date
                }
                
                work_order_item = WorkOrderItem(
                    work_order_id=local_work_order.id,
                    asset_id=asset.id,
                    asset_sn=sn,
                    asset_tag=asset.asset_tag,
                    operation_data=operation_data,
                    status="pending",
                    item_datacenter=work_order_data.datacenter
                )
                db.add(work_order_item)
            
            db.commit()
            db.refresh(local_work_order)
            print(f"[资产出入门工单] 本地工单记录已保存: ID={local_work_order.id}, 工单号={work_order_number}")
            
        except Exception as e:
            db.rollback()
            print(f"[警告] 保存本地工单记录失败: {str(e)}")
        
        # 6. 记录日志
        logger.info("资产出入门工单创建成功", extra={
            "operationObject": work_order_number or batch_id,
            "operationType": OperationType.ASSET_ENTRY_EXIT_CREATE,
            "operator": creator_name,
            "result": OperationResult.SUCCESS,
            "operationDetail": (
                f"操作内容: 建单(Create Ticket), 工单号: {work_order_number or batch_id}, "
                f"工单标题: {work_order_data.title}, 机房: {work_order_data.datacenter}, "
                f"出入类型: {work_order_data.entry_exit_type}, 出入范围: {work_order_data.entry_exit_scope}, "
                f"出入日期: {work_order_data.entry_exit_date}"
            )
        })
        
        # 7. 构建响应数据
        response_data = AssetEntryExitWorkOrderResponse(
            work_order_number=work_order_number,
            batch_id=batch_id,
            title=work_order_data.title,
            datacenter=work_order_data.datacenter,
            priority=str(work_order_data.priority),
            business_type=str(work_order_data.business_type),
            device_sns=work_order_data.device_sns,
            device_count=len(work_order_data.device_sns),
            service_content=work_order_data.service_content,
            entry_exit_type=str(work_order_data.entry_exit_type),
            entry_exit_scope=str(work_order_data.entry_exit_scope),
            entry_exit_reason=work_order_data.entry_exit_reason,
            entry_exit_date=work_order_data.entry_exit_date,
            assignee=work_order_data.assignee,
            creator_name=work_order_data.creator_name,
            status="pending",
            work_order_status="processing",  # 创建成功即为进行中
            remark=work_order_data.remark,
            attachments=work_order_data.attachments,
            source_order_number=work_order_data.source_order_number,
            sign_date=work_order_data.sign_date,
            created_at=local_work_order.created_at if local_work_order else datetime.now()
        )
        
        return ApiResponse(
            code=ResponseCode.SUCCESS,
            message="工单创建成功",
            data=response_data.dict()
        )
        
    except ValueError as e:
        logger.error("资产出入门工单创建参数验证失败", extra={
            "operationObject": work_order_data.title if 'work_order_data' in locals() else "未知工单",
            "operationType": OperationType.ASSET_ENTRY_EXIT_CREATE,
            "operator": work_order_data.creator_name if 'work_order_data' in locals() and work_order_data.creator_name else "system",
            "result": OperationResult.FAILED,
            "operationDetail": f"参数验证失败: {str(e)}"
        })
        return ApiResponse(
            code=ResponseCode.PARAM_ERROR,
            message=f"参数验证失败: {str(e)}",
            data=None
        )
    except Exception as e:
        db.rollback()
        logger.error("资产出入门工单创建系统异常", extra={
            "operationObject": work_order_data.title if 'work_order_data' in locals() else "未知工单",
            "operationType": OperationType.ASSET_ENTRY_EXIT_CREATE,
            "operator": work_order_data.creator_name if 'work_order_data' in locals() and work_order_data.creator_name else "system",
            "result": OperationResult.FAILED,
            "operationDetail": f"系统异常: {str(e)}"
        })
        return ApiResponse(
            code=ResponseCode.INTERNAL_ERROR,
            message=f"创建工单失败: {str(e)}",
            data=None
        )



@router.get("/query", summary="查询资产出入门工单",
            response_model=ApiResponse,
            responses={
                200: {
                    "description": "查询成功",
                    "content": {
                        "application/json": {
                            "example": {
                                "code": 0,
                                "message": "查询成功",
                                "data": {
                                    "work_orders": [
                                        {
                                            "batch_id": "AEE_20251210120000",
                                            "work_order_number": "WO202512101234",
                                            "arrival_order_number": None,
                                            "source_order_number": None,
                                            "operation_type": "asset_accounting",
                                            "title": "服务器设备搬入",
                                            "description": "出入类型: 搬入\n出入范围: 出入机房\n...",
                                            "status": "pending",
                                            "work_order_status": "processing",
                                            "is_timeout": False,
                                            "creator": "李四",
                                            "operator": None,
                                            "assignee": "张三",
                                            "datacenter": "DC01",
                                            "campus": None,
                                            "room": None,
                                            "created_at": "2025-12-10T12:00:00",
                                            "completed_time": None,
                                            "close_time": None,
                                            "device_count": 2,
                                            "remark": "请提前准备好机柜空间",
                                            "priority": "normal",
                                            "business_type": "other",
                                            "device_sns": ["SN123456", "SN789012"],
                                            "service_content": "新采购服务器搬入机房上架",
                                            "entry_exit_type": "move_in",
                                            "entry_exit_scope": "datacenter",
                                            "entry_exit_reason": "新设备采购到货",
                                            "entry_exit_date": "2025-12-10",
                                            "sign_date": "2025-12-10",
                                            "attachments": ["https://example.com/file1.pdf"],
                                            "creator_name": "李四",
                                            "updated_assets_count": None
                                        }
                                    ],
                                    "total": 1,
                                    "page": 1,
                                    "size": 10,
                                    "pages": 1
                                }
                            }
                        }
                    }
                }
            })
async def query_asset_entry_exit_work_orders(
    work_order_number: Optional[str] = Query(None, description="工单号"),
    batch_id: Optional[str] = Query(None, description="批次ID"),
    datacenter: Optional[str] = Query(None, description="机房"),
    priority: Optional[str] = Query(None, description="优先级"),
    business_type: Optional[str] = Query(None, description="业务类型"),
    status: Optional[str] = Query(None, description="状态"),
    assignee: Optional[str] = Query(None, description="指派人"),
    device_sn: Optional[str] = Query(None, description="设备SN"),
    entry_exit_type: Optional[str] = Query(None, description="出入类型"),
    entry_exit_scope: Optional[str] = Query(None, description="出入范围"),
    creator_name: Optional[str] = Query(None, description="创建人"),
    created_from: Optional[datetime] = Query(None, description="创建时间起始"),
    created_to: Optional[datetime] = Query(None, description="创建时间结束"),
    page: int = Query(1, ge=1, description="页码"),
    size: int = Query(10, ge=1, le=100, description="每页数量"),
    db: Session = Depends(get_db)
):
    """
    查询资产出入门工单列表
    
    ## 查询参数
    - **work_order_number**: 工单号（精确匹配）
    - **batch_id**: 批次ID（精确匹配）
    - **datacenter**: 机房（模糊匹配）
    - **priority**: 优先级（精确匹配）- normal/urgent
    - **business_type**: 业务类型（精确匹配）
    - **status**: 状态（精确匹配）- pending/processing/completed
    - **assignee**: 指派人（模糊匹配）
    - **device_sn**: 设备SN（通过工单明细关联查询）
    - **entry_exit_type**: 出入类型（精确匹配）- move_in/move_out
    - **entry_exit_scope**: 出入范围（精确匹配）- datacenter/campus/internal
    - **creator_name**: 创建人（模糊匹配）
    - **created_from/created_to**: 创建时间范围
    """
    try:
        query = db.query(WorkOrder).filter(
            WorkOrder.operation_type == "asset_accounting"
        )
        
        if work_order_number:
            query = query.filter(WorkOrder.work_order_number == work_order_number)
        
        if batch_id:
            query = query.filter(WorkOrder.batch_id == batch_id)
        
        if datacenter:
            query = query.filter(WorkOrder.datacenter.like(f"%{datacenter}%"))
        
        if status:
            query = query.filter(WorkOrder.status == status)
        
        if assignee:
            query = query.filter(WorkOrder.assignee.like(f"%{assignee}%"))
        
        if created_from:
            query = query.filter(WorkOrder.created_at >= created_from)
        
        if created_to:
            query = query.filter(WorkOrder.created_at <= created_to)
        
        # 通过extra字段过滤
        if priority:
            query = query.filter(cast(WorkOrder.extra['priority'], String) == f'"{priority}"')
        
        if business_type:
            query = query.filter(cast(WorkOrder.extra['business_type'], String) == f'"{business_type}"')
        
        if entry_exit_type:
            query = query.filter(cast(WorkOrder.extra['entry_exit_type'], String) == f'"{entry_exit_type}"')
        
        if entry_exit_scope:
            query = query.filter(cast(WorkOrder.extra['entry_exit_scope'], String) == f'"{entry_exit_scope}"')
        
        if creator_name:
            query = query.filter(cast(WorkOrder.extra['creator_name'], String).like(f'%{creator_name}%'))
        
        # 通过设备SN查询
        if device_sn:
            work_order_ids = db.query(WorkOrderItem.work_order_id).filter(
                WorkOrderItem.asset_sn == device_sn
            ).subquery()
            query = query.filter(WorkOrder.id.in_(work_order_ids))
        
        # 统计总数
        total = query.count()
        
        # 分页
        pages = (total + size - 1) // size
        work_orders = query.order_by(WorkOrder.created_at.desc()).offset((page - 1) * size).limit(size).all()
        
        # 构建响应
        work_order_list = []
        for wo in work_orders:
            extra = wo.extra or {}
            work_order_list.append({
                # ===== 核心标识 =====
                "batch_id": wo.batch_id,
                "work_order_number": wo.work_order_number,
                "arrival_order_number": wo.arrival_order_number,
                "source_order_number": wo.source_order_number,
                
                # ===== 业务信息 =====
                "operation_type": wo.operation_type,
                "title": wo.title,
                "description": wo.description,
                
                # ===== 状态管理 =====
                "status": wo.status,
                "work_order_status": wo.work_order_status or "processing",
                "is_timeout": wo.is_timeout,
                
                # ===== 人员信息 =====
                "creator": wo.creator,
                "operator": wo.operator,
                "assignee": wo.assignee,
                
                # ===== 位置信息 =====
                "datacenter": wo.datacenter,
                "campus": wo.campus,
                "room": wo.room,
                
                # ===== 时间信息 =====
                "created_at": wo.created_at.isoformat() if wo.created_at else None,
                "completed_time": wo.completed_time.isoformat() if wo.completed_time else None,
                "close_time": wo.close_time.isoformat() if wo.close_time else None,
                
                # ===== 统计信息 =====
                "device_count": wo.device_count or 0,
                
                # ===== 备注信息 =====
                "remark": wo.remark,
                
                # ===== 资产出入门专用字段 =====
                "priority": extra.get("priority"),
                "business_type": extra.get("business_type"),
                "device_sns": extra.get("device_sns", []),
                "service_content": extra.get("service_content"),
                "entry_exit_type": extra.get("entry_exit_type"),
                "entry_exit_scope": extra.get("entry_exit_scope"),
                "entry_exit_reason": extra.get("entry_exit_reason"),
                "entry_exit_date": extra.get("entry_exit_date"),
                "attachments": extra.get("attachments", []),
                "creator_name": extra.get("creator_name"),
                "updated_assets_count": extra.get("updated_assets_count"),
            })
        
        return ApiResponse(
            code=ResponseCode.SUCCESS,
            message="查询成功",
            data={
                "work_orders": work_order_list,
                "total": total,
                "page": page,
                "size": size,
                "pages": pages
            }
        )
        
    except Exception as e:
        return ApiResponse(
            code=ResponseCode.INTERNAL_ERROR,
            message=f"查询失败: {str(e)}",
            data=None
        )


@router.get("/{batch_id}", summary="获取资产出入门工单详情",
            response_model=ApiResponse,
            responses={
                200: {
                    "description": "查询成功",
                    "content": {
                        "application/json": {
                            "example": {
                                "code": 0,
                                "message": "查询成功",
                                "data": {
                                    "batch_id": "AEE_20251210120000",
                                    "work_order_number": "WO202512101234",
                                    "arrival_order_number": None,
                                    "source_order_number": None,
                                    "operation_type": "asset_accounting",
                                    "title": "服务器设备搬入",
                                    "description": "出入类型: 搬入\n出入范围: 出入机房\n出入原因: 新设备采购到货\n...",
                                    "status": "completed",
                                    "work_order_status": "completed",
                                    "is_timeout": False,
                                    "sla_countdown": None,
                                    "creator": "李四",
                                    "operator": "张三",
                                    "assignee": "张三",
                                    "reviewer": None,
                                    "datacenter": "DC01",
                                    "campus": None,
                                    "room": None,
                                    "cabinet": None,
                                    "rack_position": None,
                                    "project_number": None,
                                    "start_time": None,
                                    "expected_completion_time": None,
                                    "completed_time": "2025-12-10T14:30:00",
                                    "close_time": None,
                                    "created_at": "2025-12-10T12:00:00",
                                    "updated_at": "2025-12-10T14:30:00",
                                    "device_count": 2,
                                    "items_count": 2,
                                    "remark": "请提前准备好机柜空间",
                                    "priority": "normal",
                                    "business_type": "other",
                                    "device_sns": ["SN123456", "SN789012"],
                                    "service_content": "新采购服务器搬入机房上架",
                                    "entry_exit_type": "move_in",
                                    "entry_exit_scope": "datacenter",
                                    "entry_exit_reason": "新设备采购到货，需搬入机房进行上架部署",
                                    "entry_exit_date": "2025-12-10",
                                    "sign_date": "2025-12-10",
                                    "attachments": ["https://example.com/attachment1.pdf"],
                                    "creator_name": "李四",
                                    "updated_assets_count": 2,
                                    "callback_remark": "设备已完成搬入",
                                    "failure_reason": None,
                                    "items": [
                                        {
                                            "asset_sn": "SN123456",
                                            "asset_tag": "ASSET-001",
                                            "status": "completed",
                                            "operation_data": {}
                                        }
                                    ]
                                }
                            }
                        }
                    }
                },
                404: {"description": "工单不存在"}
            })
async def get_asset_entry_exit_work_order(
    batch_id: str = Path(..., description="批次ID"),
    db: Session = Depends(get_db)
):
    """
    获取资产出入门工单详情
    
    ## 功能说明
    根据批次ID查询资产出入门工单的完整信息，包括所有设备明细。
    
    ## 返回字段说明
    
    ### 核心标识 (4个)
    - **batch_id**: 批次ID
    - **work_order_number**: 外部工单号
    - **arrival_order_number**: 到货单号
    - **source_order_number**: 来源单号
    
    ### 业务信息 (3个)
    - **operation_type**: 操作类型 (asset_accounting)
    - **title**: 工单标题
    - **description**: 工单描述
    
    ### 状态管理 (4个)
    - **status**: 内部状态 (pending/processing/completed/failed)
    - **work_order_status**: 外部工单状态 (processing/completed/failed)
    - **is_timeout**: 是否超时
    - **sla_countdown**: SLA倒计时（秒）
    
    ### 人员信息 (4个)
    - **creator**: 创建人
    - **operator**: 操作人
    - **assignee**: 指派人
    - **reviewer**: 审核人
    
    ### 位置信息 (5个)
    - **datacenter**: 机房
    - **campus**: 园区
    - **room**: 房间
    - **cabinet**: 机柜
    - **rack_position**: 机位
    
    ### 时间信息 (6个)
    - **start_time**: 开始时间
    - **expected_completion_time**: 期望完成时间
    - **completed_time**: 完成时间
    - **close_time**: 关闭时间
    - **created_at**: 创建时间
    - **updated_at**: 更新时间
    
    ### 统计信息 (2个)
    - **device_count**: 设备数量
    - **items_count**: 明细数量
    
    ### 资产出入门专用字段 (13个)
    - **priority**: 优先级 (normal/urgent)
    - **business_type**: 业务类型 (fault_support/change_support/other)
    - **device_sns**: 设备SN列表
    - **service_content**: 服务内容
    - **entry_exit_type**: 出入类型 (move_in/move_out)
    - **entry_exit_scope**: 出入范围 (datacenter/campus/internal)
    - **entry_exit_reason**: 出入原因
    - **entry_exit_date**: 出入日期
    - **attachments**: 附件URL列表
    - **creator_name**: 创建人姓名
    - **updated_assets_count**: 更新的资产数量（回调后）
    - **callback_remark**: 回调备注
    - **failure_reason**: 失败原因
    
    ### 设备明细
    - **items**: 设备明细列表
      - **asset_sn**: 设备序列号
      - **asset_tag**: 资产标签
      - **status**: 明细状态
      - **operation_data**: 操作数据
    """
    try:
        work_order = db.query(WorkOrder).filter(
            WorkOrder.batch_id == batch_id,
            WorkOrder.operation_type == "asset_accounting"
        ).first()
        
        if not work_order:
            return ApiResponse(
                code=ResponseCode.NOT_FOUND,
                message=f"工单不存在: {batch_id}",
                data=None
            )
        
        extra = work_order.extra or {}
        
        # 获取工单明细
        items = db.query(WorkOrderItem).filter(
            WorkOrderItem.work_order_id == work_order.id
        ).all()
        
        item_list = []
        for item in items:
            item_list.append({
                "asset_sn": item.asset_sn,
                "asset_tag": item.asset_tag,
                "status": item.status,
                "operation_data": item.operation_data
            })
        
        return ApiResponse(
            code=ResponseCode.SUCCESS,
            message="查询成功",
            data={
                # ===== 核心标识 =====
                "batch_id": work_order.batch_id,
                "work_order_number": work_order.work_order_number,
                "arrival_order_number": work_order.arrival_order_number,
                "source_order_number": work_order.source_order_number,
                
                # ===== 业务信息 =====
                "operation_type": work_order.operation_type,
                "title": work_order.title,
                "description": work_order.description,
                
                # ===== 状态管理 =====
                "status": work_order.status,
                "work_order_status": work_order.work_order_status or "processing",
                "is_timeout": work_order.is_timeout,
                "sla_countdown": work_order.sla_countdown,
                
                # ===== 人员信息 =====
                "creator": work_order.creator,
                "operator": work_order.operator,
                "assignee": work_order.assignee,
                "reviewer": work_order.reviewer,
                
                # ===== 位置信息 =====
                "datacenter": work_order.datacenter,
                "campus": work_order.campus,
                "room": work_order.room,
                "cabinet": work_order.cabinet,
                "rack_position": work_order.rack_position,
                
                # ===== 项目信息 =====
                "project_number": work_order.project_number,
                
                # ===== 时间信息 =====
                "start_time": work_order.start_time.isoformat() if work_order.start_time else None,
                "expected_completion_time": work_order.expected_completion_time.isoformat() if work_order.expected_completion_time else None,
                "completed_time": work_order.completed_time.isoformat() if work_order.completed_time else None,
                "close_time": work_order.close_time.isoformat() if work_order.close_time else None,
                "created_at": work_order.created_at.isoformat() if work_order.created_at else None,
                "updated_at": work_order.updated_at.isoformat() if work_order.updated_at else None,
                
                # ===== 统计信息 =====
                "device_count": work_order.device_count or 0,
                "items_count": len(item_list),
                
                # ===== 备注信息 =====
                "remark": work_order.remark,
                
                # ===== 资产出入门专用字段 =====
                "priority": extra.get("priority"),
                "business_type": extra.get("business_type"),
                "device_sns": extra.get("device_sns", []),
                "service_content": extra.get("service_content"),
                "entry_exit_type": extra.get("entry_exit_type"),
                "entry_exit_scope": extra.get("entry_exit_scope"),
                "entry_exit_reason": extra.get("entry_exit_reason"),
                "entry_exit_date": extra.get("entry_exit_date"),
                "attachments": extra.get("attachments", []),
                "creator_name": extra.get("creator_name"),
                "updated_assets_count": extra.get("updated_assets_count"),
                "callback_remark": extra.get("callback_remark"),
                "failure_reason": extra.get("failure_reason"),
                
                # ===== 设备明细 =====
                "items": item_list
            }
        )
        
    except Exception as e:
        return ApiResponse(
            code=ResponseCode.INTERNAL_ERROR,
            message=f"查询失败: {str(e)}",
            data=None
        )



@router.post("/callback", summary="资产出入门工单回调（已废弃，请使用 /api/v1/receiving/work-order/notify）",
             deprecated=True,
             response_model=ApiResponse,
             responses={
                 200: {
                     "description": "回调处理成功",
                     "content": {
                         "application/json": {
                             "examples": {
                                 "completed": {
                                     "summary": "工单完成",
                                     "value": {
                                         "code": 0,
                                         "message": "回调处理成功",
                                         "data": {
                                             "batch_id": "AEE_20251210120000",
                                             "work_order_status": "completed",
                                             "entry_exit_type": "move_in",
                                             "updated_assets_count": 2,
                                             "device_direction": "inbound"
                                         }
                                     }
                                 },
                                 "failed": {
                                     "summary": "工单失败",
                                     "value": {
                                         "code": 0,
                                         "message": "回调处理成功",
                                         "data": {
                                             "batch_id": "AEE_20251210120000",
                                             "work_order_status": "failed"
                                         }
                                     }
                                 }
                             }
                         }
                     }
                 },
                 404: {"description": "工单不存在"},
                 400: {"description": "无效的状态参数"}
             })
async def asset_entry_exit_callback(
    batch_id: str = Body(..., description="批次ID"),
    status: str = Body(..., description="状态：completed-完成, failed-失败"),
    operator: Optional[str] = Body(None, description="操作人"),
    remark: Optional[str] = Body(None, description="备注"),
    db: Session = Depends(get_db)
):
    """
    资产出入门工单回调接口
    
    ## ⚠️ 已废弃警告
    此接口已废弃，请使用统一的工单回调接口：
    **POST /api/v1/receiving/work-order/notify**
    
    统一接口支持所有类型工单的回调通知，包括：
    - receiving（设备到货）
    - racking（设备上架）
    - configuration（设备增配）
    - power_on（电源开启）
    - power_off（电源关闭）
    - asset_accounting（资产出入门）
    
    ## 功能说明
    外部工单系统处理完成后，通过此接口回调更新工单状态。
    工单完成时，会自动同步更新相关资产的出入库状态（device_direction字段）。
    
    ## 参数说明
    - **batch_id**: 批次ID（必填）
    - **status**: 状态（必填）
      - completed: 处理完成
      - failed: 处理失败
    - **operator**: 操作人（可选）
    - **remark**: 备注（可选）
    
    ## 状态流转
    - processing → completed（成功）
    - processing → failed（失败）
    
    ## 资产状态同步
    当工单状态为 completed 时，会自动更新资产的 device_direction 字段：
    - 搬入(move_in) → 入库(inbound)
    - 搬出(move_out) → 出库(outbound)
    
    ## 返回数据
    - **batch_id**: 批次ID
    - **work_order_status**: 工单状态
    - **entry_exit_type**: 出入类型（仅completed时返回）
    - **updated_assets_count**: 更新的资产数量（仅completed时返回）
    - **device_direction**: 资产设备去向（仅completed时返回）
    """
    try:
        # 查找工单
        work_order = db.query(WorkOrder).filter(
            WorkOrder.batch_id == batch_id,
            WorkOrder.operation_type == "asset_accounting"
        ).first()
        
        if not work_order:
            return ApiResponse(
                code=ResponseCode.NOT_FOUND,
                message=f"工单不存在: {batch_id}",
                data=None
            )
        
        # 验证状态
        if status not in ['completed', 'failed']:
            return ApiResponse(
                code=ResponseCode.PARAM_ERROR,
                message=f"无效的状态: {status}，只能是 completed 或 failed",
                data=None
            )
        
        # 更新工单状态
        work_order.work_order_status = status
        if status == 'completed':
            work_order.status = 'completed'
            work_order.completed_time = datetime.now()
            
            # 工单完成时，同步更新资产的出入库状态
            extra = work_order.extra or {}
            entry_exit_type = extra.get("entry_exit_type")
            device_sns = extra.get("device_sns", [])
            
            if entry_exit_type and device_sns:
                # 根据出入类型确定资产的设备去向
                # move_in (搬入) -> inbound (入库)
                # move_out (搬出) -> outbound (出库)
                new_direction = "inbound" if entry_exit_type == "move_in" else "outbound"
                
                # 更新所有相关资产的设备去向
                updated_count = 0
                for sn in device_sns:
                    asset = db.query(Asset).filter(Asset.serial_number == sn).first()
                    if asset:
                        old_direction = asset.device_direction
                        asset.device_direction = new_direction
                        updated_count += 1
                        
                        # 记录资产状态变更日志
                        logger.info("资产出入库状态更新", extra={
                            "operationObject": sn,
                            "operationType": "asset.device_direction_update",
                            "operator": operator or "system",
                            "result": OperationResult.SUCCESS,
                            "operationDetail": f"设备去向从 {old_direction} 更新为 {new_direction}，关联工单: {batch_id}"
                        })
                
                # 记录更新数量到extra
                extra['updated_assets_count'] = updated_count
                work_order.extra = extra
                
        elif status == 'failed':
            work_order.status = 'failed'
        
        if operator:
            work_order.operator = operator
        
        # 更新extra中的备注
        if remark:
            extra = work_order.extra or {}
            extra['callback_remark'] = remark
            work_order.extra = extra
        
        db.commit()
        
        # 记录日志
        logger.info("资产出入门工单回调成功", extra={
            "operationObject": batch_id,
            "operationType": "asset_entry_exit.callback",
            "operator": operator or "system",
            "result": OperationResult.SUCCESS,
            "operationDetail": f"工单状态更新为: {status}"
        })
        
        # 构建响应数据
        response_data = {
            "batch_id": batch_id,
            "work_order_status": status
        }
        
        # 如果是完成状态，返回更新的资产信息
        if status == 'completed':
            extra = work_order.extra or {}
            response_data["entry_exit_type"] = extra.get("entry_exit_type")
            response_data["updated_assets_count"] = extra.get("updated_assets_count", 0)
            response_data["device_direction"] = "inbound" if extra.get("entry_exit_type") == "move_in" else "outbound"
        
        return ApiResponse(
            code=ResponseCode.SUCCESS,
            message="回调处理成功",
            data=response_data
        )
        
    except Exception as e:
        db.rollback()
        logger.error("资产出入门工单回调失败", extra={
            "operationObject": batch_id,
            "operationType": "asset_entry_exit.callback",
            "operator": operator or "system",
            "result": OperationResult.FAILED,
            "operationDetail": f"回调失败: {str(e)}"
        })
        return ApiResponse(
            code=ResponseCode.INTERNAL_ERROR,
            message=f"回调处理失败: {str(e)}",
            data=None
        )



@router.get("/template/download", summary="下载资产出入门工单模板",
            responses={
                200: {
                    "description": "Excel模板文件",
                    "content": {
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {}
                    }
                },
                404: {"description": "模板文件不存在"}
            })
async def download_asset_entry_exit_template():
    """
    下载资产出入门工单Excel模板
    
    ## 功能说明
    下载用于批量创建资产出入门工单的Excel模板文件。
    
    ## 返回
    - Excel文件（.xlsx格式）
    """
    # 模板文件路径
    template_path = FilePath(__file__).parent.parent.parent.parent / "asset_door.xlsx"
    
    if not template_path.exists():
        raise HTTPException(status_code=404, detail="模板文件不存在")
    
    return FileResponse(
        path=str(template_path),
        filename="资产出入门工单模板.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@router.post("/template/upload", summary="上传并解析资产出入门工单模板",
             response_model=ApiResponse,
             responses={
                 200: {
                     "description": "解析成功",
                     "content": {
                         "application/json": {
                             "examples": {
                                 "全部验证成功": {
                                     "summary": "所有SN验证通过",
                                     "value": {
                                         "code": 0,
                                         "message": "验证成功，所有设备SN均存在",
                                         "data": {
                                             "success": True,
                                             "total": 2,
                                             "valid_count": 2,
                                             "invalid_count": 0,
                                             "valid_records": [
                                                 {
                                                     "sn": "SN001",
                                                     "datacenter": "DC01",
                                                     "device_type": "服务器",
                                                     "secondary_category": "x86服务器",
                                                     "tertiary_category": "机架式服务器",
                                                     "device_model": "R740",
                                                     "quantity": 1,
                                                     "remark": "测试设备",
                                                     "asset_id": 100,
                                                     "asset_name": "Dell服务器",
                                                     "asset_tag": "AT001"
                                                 }
                                             ],
                                             "invalid_records": []
                                         }
                                     }
                                 },
                                 "部分验证失败": {
                                     "summary": "部分SN验证失败",
                                     "value": {
                                         "code": 4000,
                                         "message": "验证完成，成功 1 条，失败 1 条",
                                         "data": {
                                             "success": False,
                                             "total": 2,
                                             "valid_count": 1,
                                             "invalid_count": 1,
                                             "valid_records": [
                                                 {
                                                     "sn": "SN001",
                                                     "datacenter": "DC01",
                                                     "device_type": "服务器",
                                                     "asset_id": 100,
                                                     "asset_name": "Dell服务器",
                                                     "asset_tag": "AT001"
                                                 }
                                             ],
                                             "invalid_records": [
                                                 {"row": 3, "sn": "SN999", "error": "设备SN不存在"}
                                             ]
                                         }
                                     }
                                 }
                             }
                         }
                     }
                 }
             })
async def upload_asset_entry_exit_template(
    file: UploadFile = File(..., description="Excel模板文件（.xlsx格式）"),
    db: Session = Depends(get_db)
):
    """
    上传并解析资产出入门工单Excel模板
    
    ## 功能说明
    上传填写好的Excel模板，解析内容并验证设备SN是否存在于资产库中。
    
    ## 模板列说明
    | 列名 | 说明 | 是否必填 |
    |------|------|----------|
    | SN | 设备序列号 | 是 |
    | IDC机房 | 机房名称 | 否 |
    | 设备类型 | 设备类型 | 否 |
    | 二级分类 | 二级分类 | 否 |
    | 三级分类 | 三级分类 | 否 |
    | 设备机型 | 设备机型 | 否 |
    | 数量 | 数量 | 否 |
    | 备注 | 备注信息 | 否 |
    
    ## 返回数据
    - **success**: 是否全部验证成功
    - **total**: 总记录数
    - **valid_count**: 有效记录数（SN存在）
    - **invalid_count**: 无效记录数（SN不存在）
    - **valid_records**: 验证通过的完整记录列表
      - sn: 设备SN
      - datacenter: IDC机房
      - device_type: 设备类型
      - secondary_category: 二级分类
      - tertiary_category: 三级分类
      - device_model: 设备机型
      - quantity: 数量
      - remark: 备注
      - asset_id: 资产ID（系统匹配）
      - asset_name: 资产名称（系统匹配）
      - asset_tag: 资产标签（系统匹配）
    - **invalid_records**: 验证失败的记录列表
      - row: 行号
      - sn: 设备SN
      - error: 错误原因
    """
    try:
        # 1. 验证文件类型
        if not file.filename.endswith('.xlsx'):
            return ApiResponse(
                code=ResponseCode.BAD_REQUEST,
                message="只支持.xlsx格式的Excel文件",
                data=None
            )
        
        # 2. 读取文件内容
        content = await file.read()
        
        # 3. 解析Excel
        try:
            import openpyxl
            from io import BytesIO
            
            wb = openpyxl.load_workbook(BytesIO(content))
            ws = wb.active
        except Exception as e:
            return ApiResponse(
                code=ResponseCode.BAD_REQUEST,
                message=f"Excel文件解析失败: {str(e)}",
                data=None
            )
        
        # 4. 获取表头
        headers = [cell.value for cell in ws[1]]
        
        # 验证必要的列
        if 'SN' not in headers:
            return ApiResponse(
                code=ResponseCode.BAD_REQUEST,
                message="Excel文件缺少必要的列: SN",
                data=None
            )
        
        # 列索引映射
        col_map = {header: idx for idx, header in enumerate(headers) if header}
        
        # 5. 解析数据行
        valid_records = []
        invalid_records = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 跳过空行
            if not any(row):
                continue
            
            # 获取SN
            sn = row[col_map.get('SN', 0)]
            if not sn:
                invalid_records.append({
                    "row": row_idx,
                    "sn": None,
                    "error": "SN为空"
                })
                continue
            
            sn = str(sn).strip()
            
            # 构建记录
            record = {
                "sn": sn,
                "datacenter": str(row[col_map.get('IDC机房', 1)] or '').strip() if col_map.get('IDC机房') is not None else None,
                "device_type": str(row[col_map.get('设备类型', 2)] or '').strip() if col_map.get('设备类型') is not None else None,
                "secondary_category": str(row[col_map.get('二级分类', 3)] or '').strip() if col_map.get('二级分类') is not None else None,
                "tertiary_category": str(row[col_map.get('三级分类', 4)] or '').strip() if col_map.get('三级分类') is not None else None,
                "device_model": str(row[col_map.get('设备机型', 5)] or '').strip() if col_map.get('设备机型') is not None else None,
                "quantity": row[col_map.get('数量', 6)] if col_map.get('数量') is not None else None,
                "remark": str(row[col_map.get('备注', 7)] or '').strip() if col_map.get('备注') is not None else None,
            }
            
            # 验证SN是否存在
            asset = db.query(Asset).filter(Asset.serial_number == sn).first()
            
            if asset:
                # SN存在，添加资产信息
                record["asset_id"] = asset.id
                record["asset_name"] = asset.name
                record["asset_tag"] = asset.asset_tag
                # 获取分类名称
                record["category"] = asset.category.name if asset.category else None
                valid_records.append(record)
            else:
                # SN不存在
                invalid_records.append({
                    "row": row_idx,
                    "sn": sn,
                    "error": "设备SN不存在"
                })
        
        # 6. 返回结果
        total = len(valid_records) + len(invalid_records)
        
        # 全部验证成功
        if len(invalid_records) == 0 and len(valid_records) > 0:
            return ApiResponse(
                code=ResponseCode.SUCCESS,
                message="验证成功，所有设备SN均存在",
                data={
                    "success": True,
                    "total": total,
                    "valid_count": total,
                    "invalid_count": 0,
                    "valid_records": valid_records,
                    "invalid_records": []
                }
            )
        
        # 有验证失败的记录
        return ApiResponse(
            code=ResponseCode.BAD_REQUEST if invalid_records else ResponseCode.SUCCESS,
            message=f"验证完成，成功 {len(valid_records)} 条，失败 {len(invalid_records)} 条",
            data={
                "success": False,
                "total": total,
                "valid_count": len(valid_records),
                "invalid_count": len(invalid_records),
                "valid_records": valid_records,
                "invalid_records": invalid_records
            }
        )
        
    except Exception as e:
        return ApiResponse(
            code=ResponseCode.INTERNAL_ERROR,
            message=f"解析失败: {str(e)}",
            data=None
        )
